import numpy as np
import openpyxl as xl
import pandas as pd
import os
from decimal import Decimal,ROUND_HALF_UP
crwd = os.getcwd()

def round(num):
    num = str(num)
    num = Decimal(num).quantize(Decimal('0.01'),ROUND_HALF_UP)
    return num

def parentDir():
    cwd = os.getcwd()
    os.chdir('..')
    ucwd =os.getcwd()
    os.chdir(cwd)
    return ucwd

def find(sheet,val):
    row_max = sheet.max_row
    i = 1
    row_index = 0
    for rows in sheet.iter_rows(min_row=1,max_row=row_max,min_col=1,max_col=1):
        for cell in rows:
            if(cell.value==val):
                row_index = i
                print("exists")
                break
            i = i + 1
    if(row_index == 0):
        print("not exists")
        row_index = i
    return  row_index

def add_data(sheet,row,val,rcs):
    i = 0
    sheet.cell(row=row,column=i+1).value = val
    for cols in sheet.iter_cols(min_row=row,max_row=row,min_col=2,max_col=1+len(rcs)):
        for cell in cols:
            sheet.cell(row=row,column=i+2).value = rcs[i]
            i = i + 1
    
def show(sheet):
    for rows in sheet.iter_cols(min_row=1,max_row=8,min_col=2,max_col=5):
        for cell in rows:
            print(cell.value)
        print()
    
def sort(xl_name,tag):
    print("sort data")
    df = pd.read_excel(xl_name,sheet_name="directivity")
    df = df.sort_values(tag,ascending=True)
    df.to_excel(xl_name,index=False)
     
def init(path,xl_name,tag):
    if not (os.path.exists(path)):
        print("make excel sheet")
        os.chdir(parentDir())
        book = xl.Workbook()
        sheet= book.active
        sheet.title = "directivity"
        row = 'A1'
        ang = np.zeros([73])
        for p in range(0,73):
            ang[p] = 5*p 
            
        i=0
        sheet.cell(row=1,column=1).value = tag
        for cols in sheet.iter_cols(min_row=1,max_row=1,min_col=2,max_col=2+72):
            for cell in cols:
                cell.value = ang[i]
                i = i + 1
        book.save(xl_name)
    else:
        print("load excel sheet")
        os.chdir(parentDir())
        book = xl.load_workbook(path)
    return book
